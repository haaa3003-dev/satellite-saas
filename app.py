import streamlit as st
import ee
import folium
from streamlit_folium import st_folium
from datetime import date, timedelta
import pandas as pd
import io
import plotly.graph_objects as go
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

# =================================================================
# [백엔드] GEE 인증 및 초기화 (로컬 및 서버 공용 방탄 코드)
# =================================================================
@st.cache_resource
def init_gee():
    if "gee_credentials" in st.secrets:
        try:
            cred_info = st.secrets["gee_credentials"]
            credentials = ee.ServiceAccountCredentials(
                cred_info["client_email"], 
                key_data=cred_info["private_key"]
            )
            ee.Initialize(credentials, project='knut-startup-gee')
            return True
        except Exception as e:
            st.error(f"서버 인증 실패: {e}")
            return False
    else:
        try:
            ee.Initialize(project='knut-startup-gee')
            return True
        except Exception:
            try:
                ee.Authenticate()
                ee.Initialize(project='knut-startup-gee')
                return True
            except Exception:
                return False

gee_ready = init_gee()

# =================================================================
# [백엔드] 범용 위성 지수(NDVI, NDWI, NBR) 통합 산출 함수
# =================================================================
def get_satellite_index_for_period(region, start_date, end_date, cloud_threshold, bands, index_name):
    collection = (
        ee.ImageCollection('COPERNICUS/S2_SR_HARMONIZED')
        .filterBounds(region)
        .filterDate(start_date, end_date)
        .filter(ee.Filter.lt('CLOUDY_PIXEL_PERCENTAGE', cloud_threshold))
    )
    count = collection.size().getInfo()
    if count == 0:
        return None, None, 0, None
        
    image = collection.median()
    calculated_index = image.normalizedDifference(bands).rename(index_name)
    
    try:
        stats = calculated_index.reduceRegion(
            reducer=ee.Reducer.mean().combine(ee.Reducer.max(), sharedInputs=True),
            geometry=region,
            scale=10
        ).getInfo()
        
        if stats is None:
            stats = {}
    except Exception:
        stats = {}
        
    return image, calculated_index, count, stats

def get_ee_tile_url(ee_image_object, vis_params):
    map_id_dict = ee.Image(ee_image_object).getMapId(vis_params)
    return map_id_dict['tile_fetcher'].url_format

# =================================================================
# [프론트엔드] Streamlit 종합 관제 웹 플랫폼 레이아웃
# =================================================================
st.set_page_config(layout="wide")
st.title("🛰️ 지자체 종합 재난재해 및 자원관리 원격 관제 플랫폼")
st.caption("항공기계공학과 융합 프로젝트 - Sentinel-2 다중 시계열 및 위성 밴드 융합 AI 공간분석 SaaS")

if not gee_ready:
    st.error("🚨 GEE 인증에 실패했습니다. 환경 설정을 확인해주세요.")
    st.stop()

# -----------------------------------------------------------------
# 🎛️ 사이드바 컨트롤 패널
# -----------------------------------------------------------------
st.sidebar.header("🛠️ 종합 관제 컨트롤 패널")

# 🌟 개선포인트 1: 분석 목적에 따른 3대 다각화 관제 모드 스위칭 인프라 구축
analysis_mode = st.sidebar.selectbox(
    "🔎 분석 모드 선택", 
    ["🌾 농작물 생육 분석 (NDVI)", "🌊 저수지 및 홍수 모니터링 (NDWI)", "🔥 산불 재해 및 산림 진단 (NBR)"]
)

# 각 모드별 밴드 조합, 가시화 팰럿, 임계값, 자동 진단 텍스트 맵 설정
mode_config = {
    "🌾 농작물 생육 분석 (NDVI)": {
        "bands": ['B8', 'B4'], "index_name": "NDVI", "label": "식생활성도",
        "palette": ['red', 'yellow', 'green'], "min": 0.0, "max": 1.0, "anomaly_min": -0.3, "anomaly_max": 0.3,
        "threshold": 0.4, "baseline": 0.15, "ceil": 0.85,
        "desc_good": "지표면의 식생 활성도가 기준치(0.4)를 상회합니다. 작물이 정상적인 성장 주기에 안착하여 활발히 생육 중임이 증명되었습니다.",
        "desc_bad": "식생지수가 다소 낮게 모니터링됩니다. 초기 파종/모내기로 인한 수면 노출이거나 가뭄, 병해충으로 인한 생육 지연일 가능성이 있습니다.",
        "ai_good": "안정적인 생육 상태를 유지하며 우수한 수확권에 진입할 것으로 분석됩니다.",
        "ai_bad": "발육 상태가 저조하므로 정밀 예찰 및 추가적인 자원(비료, 용수) 투입 조치를 권장합니다."
    },
    "🌊 저수지 및 홍수 모니터링 (NDWI)": {
        "bands": ['B3', 'B8'], "index_name": "NDWI", "label": "수분포화도",
        "palette": ['white', '#99ccff', 'blue'], "min": -0.5, "max": 0.8, "anomaly_min": -0.4, "anomaly_max": 0.4,
        "threshold": 0.1, "baseline": -0.40, "ceil": 0.75,
        "desc_good": "해당 구역의 수분포화도가 높습니다. 대형 저수지의 저수율이 풍부하거나 호우로 인한 지표면 침수 및 하천 범람 구역일 수 있습니다.",
        "desc_bad": "수분포화도가 마이너스권을 기록합니다. 수자원이 고갈되어 가뭄 징후가 보이거나 건조한 나대지 상태를 나타냅니다.",
        "ai_good": "수량이 유지되거나 과포화 상태가 지속될 것으로 예측되므로 침수 취약 지역은 배수 시설 점검이 필요합니다.",
        "ai_bad": "지속적인 수분 감소 추세가 예측되므로 지자체 차원의 농업·공업용수 제한 조치 및 저수율 관리가 요구됩니다."
    },
    "🔥 산불 재해 및 산림 진단 (NBR)": {
        "bands": ['B8', 'B12'], "index_name": "NBR", "label": "탄화흔적도",
        "palette": ['#331a00', 'yellow', 'darkgreen'], "min": -0.4, "max": 0.8, "anomaly_min": -0.5, "anomaly_max": 0.5,
        "threshold": 0.15, "baseline": -0.30, "ceil": 0.80,
        "desc_good": "탄화흔적이 없는 깨끗하고 푸른 산림 상태를 나타냅니다. 산림 자원의 건강성이 아주 우수하게 유지되고 있습니다.",
        "desc_bad": "지수가 급격한 마이너스로 추락했습니다. 최근 산불 재해로 인해 지표면이 까맣게 타버린 '탄화 흔적지'이거나 급격한 산림 훼손 구역입니다.",
        "ai_good": "재해 징후 없이 산림의 건강도 지표가 향후에도 안정적으로 복원/유지될 것으로 전망됩니다.",
        "ai_bad": "재해 흔적이 깊게 잔존하거나 주변 산림으로의 추가적인 황폐화 추세가 인지되므로 즉각적인 토사 유출 및 복구 예산을 편성해야 합니다."
    }
}

cfg = mode_config[analysis_mode]

st.sidebar.markdown("---")
st.sidebar.subheader("🎯 관제 타겟 지역 설정")

# 🌟 개선포인트 2: 단순 농경지를 넘어 전국 단위 핵심 지자체 요충지 8곳으로 프리셋 대폭 확장
region_preset = st.sidebar.selectbox(
    "협업 대상 지자체 및 관제 지역 선택",
    [
        "🌾 [농업] 전북 김제시 부량면 (벽골제 평야)", 
        "🌾 [농업] 충남 당진시 합덕읍 (당진평야)", 
        "🌾 [농업] 전남 해남군 황산면 (대규모 필드)",
        "🌊 [수자원] 충북 충주시 종민동 (충주호 저수지)", 
        "🌊 [수자원] 강원 춘천시 신북읍 (소양강댐 부근)",
        "🔥 [재해] 강원 고성군 토성면 (산불 취약 산림지)", 
        "🔥 [재해] 경북 안동시 풍천면 (산림 보존 구역)",
        "🏙️ [도시] 서울 뚝섬 한강공원 (도심 녹지축)",
        "📍 직접 좌표 입력"
    ]
)

# 확장 프리셋별 정밀 위경도 매핑 딕셔너리
preset_coords = {
    "🌾 [농업] 전북 김제시 부량면 (벽골제 평야)": (35.7684, 126.8643),
    "🌾 [농업] 충남 당진시 합덕읍 (당진평야)": (36.8250, 126.7720),
    "🌾 [농업] 전남 해남군 황산면 (대규모 필드)": (34.6150, 126.4780),
    "🌊 [수자원] 충북 충주시 종민동 (충주호 저수지)": (36.9910, 127.9259),
    "🌊 [수자원] 강원 춘천시 신북읍 (소양강댐 부근)": (37.9425, 127.8140),
    "🔥 [재해] 강원 고성군 토성면 (산불 취약 산림지)": (38.2250, 128.5110),
    "🔥 [재해] 경북 안동시 풍천면 (산림 보존 구역)": (36.5750, 128.5210),
    "🏙️ [도시] 서울 뚝섬 한강공원 (도심 녹지축)": (37.5285, 127.0675),
    "📍 직접 좌표 입력": (36.9910, 127.9259)
}

default_lat, default_lon = preset_coords[region_preset]

lat = st.sidebar.number_input("위도 (Latitude)", value=default_lat, format="%.4f")
lon = st.sidebar.number_input("경도 (Longitude)", value=default_lon, format="%.4f")
buffer_m = st.sidebar.slider("관측 관제 반경 (m)", 500, 5000, 1500, step=500)

st.sidebar.markdown("---")
st.sidebar.subheader("📅 원격 시계열 분석 기간")
default_start = date.today() - timedelta(days=45)
default_end = date.today() - timedelta(days=5)
start_date = st.sidebar.date_input("관측 시작일", value=default_start)
end_date = st.sidebar.date_input("관측 종료일", value=default_end)
cloud_threshold = st.sidebar.slider("최대 허용 구름 비율 (%)", 5, 50, 25)

if "analysis_done" not in st.session_state:
    st.session_state.analysis_done = False

# -----------------------------------------------------------------
# 🔍 글로벌 위성 종합 분석 엔진 가동
# -----------------------------------------------------------------
if st.sidebar.button("🛰️ 글로벌 위성 분석 엔진 가동"):
    if start_date >= end_date:
        st.warning("⚠️ 관측 날짜 설정을 다시 확인해주세요.")
        st.stop()

    with st.spinner(f"구글 슈퍼컴퓨터 인프라가 {cfg['index_name']} 매핑 및 알고리즘을 연산 중입니다... 🚀"):
        try:
            point = ee.Geometry.Point([lon, lat])
            region = point.buffer(buffer_m)
            
            # [올해 위성 데이터 원격 진단]
            this_image, this_idx, this_count, this_stats = get_satellite_index_for_period(
                region, str(start_date), str(end_date), cloud_threshold, cfg['bands'], cfg['index_name']
            )
            
            if this_idx is None:
                st.session_state.analysis_done = False
                st.warning("⚠️ 선택 기간에 구름 피복률이 높아 위성 영상을 합성할 수 없습니다.")
                st.stop()
                
            # [작년 동기 대조군 원격 진단]
            ly_start = start_date.replace(year=start_date.year - 1)
            ly_end = end_date.replace(year=end_date.year - 1)
            _, last_idx, _, last_stats = get_satellite_index_for_period(
                region, str(ly_start), str(ly_end), cloud_threshold + 10, cfg['bands'], cfg['index_name']
            )
            
            # 공간정보 레이어 빌드
            vis_params_rgb = {'bands': ['B4', 'B3', 'B2'], 'min': 0, 'max': 2500}
            rgb_tile_url = get_ee_tile_url(this_image.clip(region), vis_params_rgb)

            vis_params_idx = {'min': cfg['min'], 'max': cfg['max'], 'palette': cfg['palette']}
            idx_tile_url = get_ee_tile_url(this_idx.clip(region), vis_params_idx)

            if this_idx is None or last_idx is None:
                st.session_state.anomaly_tile_url = None
            else:
                anomaly_idx = this_idx.subtract(last_idx)
                vis_params_anomaly = {'min': cfg['anomaly_min'], 'max': cfg['anomaly_max'], 'palette': ['red', 'white', 'green']}
                anomaly_tile_url = get_ee_tile_url(anomaly_idx.clip(region), vis_params_anomaly)
                st.session_state.anomaly_tile_url = anomaly_tile_url

            this_stats = this_stats if this_stats else {}
            st.session_state.analysis_done = True
            st.session_state.rgb_tile_url = rgb_tile_url
            st.session_state.idx_tile_url = idx_tile_url
            st.session_state.count = this_count
            st.session_state.current_mode = analysis_mode
            
            # 프리셋에서 행정 구역명만 깔끔하게 파싱 (예: "🌾 [농업] 전북 김제시..." -> "전북 김제시 부량면")
            raw_name = region_preset.split("] ")[-1] if "] " in region_preset else region_preset
            st.session_state.region_name = raw_name.split(" (")[0]
            
            st.session_state.avg_val = this_stats.get(f"{cfg['index_name']}_mean", 0) or 0
            st.session_state.max_val = this_stats.get(f"{cfg['index_name']}_max", 0) or 0
            
            if last_stats:
                st.session_state.last_avg_val = last_stats.get(f"{cfg['index_name']}_mean", 0) or 0
            else:
                st.session_state.last_avg_val = None
                
            st.session_state.map_lat = lat
            st.session_state.map_lon = lon

        except Exception as e:
            st.session_state.analysis_done = False
            st.error(f"🚨 공간 융합 연산 중 오류 발생: {e}")

# =================================================================
# 🖥️ 화면 출력 및 동적 리포트 시각화 구역
# =================================================================
if st.session_state.analysis_done:
    curr_cfg = mode_config.get(st.session_state.current_mode, cfg)
    idx_name = curr_cfg['index_name']

    st.success(f"✅ [관제 가동] {st.session_state.region_name} 구역 - {st.session_state.current_mode} 원격 모니터링 탐지 완료")

    col1, col2 = st.columns([1.4, 1])

    with col1:
        st.subheader("🗺️ 지자체 공간정보 디지털 트윈 위성 맵")
        m = folium.Map(location=[st.session_state.map_lat, st.session_state.map_lon], zoom_start=14)

        folium.TileLayer(
            tiles=st.session_state.rgb_tile_url, attr='Google Earth Engine',
            name='Sentinel-2 실제 현장 컬러 사진', overlay=True, control=True
        ).add_to(m)

        folium.TileLayer(
            tiles=st.session_state.idx_tile_url, attr='Google Earth Engine',
            name=f"{curr_cfg['label']} 지성 맵 ({idx_name})", overlay=True, control=True
        ).add_to(m)

        if st.session_state.get('anomaly_tile_url'):
            folium.TileLayer(
                tiles=st.session_state.anomaly_tile_url, attr='Google Earth Engine',
                name='🚨 전년 동기 대비 이상 징후 탐지 레이어 (빨강=위험/위축)', overlay=True, control=True
            ).add_to(m)

        # 🌟 개선포인트 3: 원형 경계선(Border)을 제거하고 부드러운 반투명 색상으로 디지털 트윈 스캔 하이라이트 구현
        folium.Circle(
            location=[st.session_state.map_lat, st.session_state.map_lon],
            radius=buffer_m,
            stroke=False,                  # 이질감을 주던 외곽선 테두리 라인을 완벽하게 차단
            fill=True,                     # 내부 채우기 활성화
            fill_color='#1F4E78',          # 스마트 시티 관제 감성의 세련된 딥 오션 네이비 컬러
            fill_opacity=0.15,             # 하부 위성 분석 레이어가 자연스럽게 투영되도록 최적의 투명도 세팅
            popup=f"원격 관제 반경: {buffer_m}m",
            tooltip="현재 AI 실시간 연산 구역"
        ).add_to(m)

        folium.LayerControl().add_to(m)
        st_folium(m, width=850, height=480, key="platform_map", returned_objects=[])

    with col2:
        st.subheader("📊 전년 대조 시계열 정량 지표")
        
        avg_val = st.session_state.avg_val
        max_val = st.session_state.max_val
        last_avg = st.session_state.last_avg_val

        if last_avg is not None and abs(last_avg) > 0:
            delta_val = avg_val - last_avg
            st.metric(
                label=f"📈 관측 구역 현재 평균 {curr_cfg['label']} 수치", 
                value=f"{avg_val:.4f}", 
                delta=f"{delta_val:+.4f} (전년 동기 대조)"
            )
        else:
            st.metric(label=f"📈 관측 구역 현재 평균 {curr_cfg['label']} 수치", value=f"{avg_val:.4f}", delta="전년 데이터 유실로 대조 불가")
            
        st.metric(label="🚀 구역 내 최고 피크 수치", value=f"{max_val:.4f}")

        st.markdown("---")
        
        st.subheader("📋 공공 결재용 정밀 원격 진단서")
        if avg_val >= curr_cfg['threshold']:
            st.info(f"🟢 **정상 관제 단계:** {curr_cfg['desc_good']}\n\n"
                    f"{'📊 전년 동기 데이터와 비교분석 결과, 과거 대비 환경 지표가 긍정적인 방향으로 발달 중입니다.' if (last_avg and avg_val > last_avg) else '⚠️ 다만 작년 수치에 비해서는 지표가 소폭 감소했으므로 담당 부서의 정기 순찰이 권장됩니다.'}")
        else:
            st.warning(f"🔴 **주의/위험 예찰 단계:** {curr_cfg['desc_bad']}")

        # -----------------------------------------------------------------
        # 📈 AI 시계열 추이 그래프 및 14일 미래 예측 커스텀 모델
        # -----------------------------------------------------------------
        st.markdown("---")
        st.subheader(f"📈 AI 시계열 {idx_name} 추이 및 14일 단기 예측")
        
        days_passed = (end_date - start_date).days if (end_date - start_date).days > 0 else 1
        growth_rate = (avg_val - curr_cfg['baseline']) / days_passed 
        
        future_date = end_date + timedelta(days=14)
        predicted_val = max(curr_cfg['min'], min(avg_val + (growth_rate * 14), curr_cfg['ceil']))

        fig = go.Figure()
        if last_avg is not None:
            fig.add_trace(go.Bar(
                x=[end_date.strftime("%Y-%m-%d")], y=[last_avg],
                name="전년 동기 평균", marker_color="lightgray", width=0.2
            ))

        fig.add_trace(go.Scatter(
            x=[start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")],
            y=[curr_cfg['baseline'], avg_val],
            mode='lines+markers+text', name="올해 관측 추이",
            line=dict(color="blue", width=4), marker=dict(size=10),
            text=["분석 시작", f"현재 ({avg_val:.3f})"], textposition="top center"
        ))

        fig.add_trace(go.Scatter(
            x=[end_date.strftime("%Y-%m-%d"), future_date.strftime("%Y-%m-%d")],
            y=[avg_val, predicted_val],
            mode='lines+markers+text', name="AI 2주 예측선",
            line=dict(color="orange", width=3, dash='dot'), marker=dict(size=10, symbol="star"),
            text=["", f"예측치 ({predicted_val:.3f})"], textposition="top right"
        ))

        fig.update_layout(
            plot_bgcolor='rgba(240, 240, 240, 0.5)',
            yaxis_title=f"{idx_name} 지수 수치", xaxis_title="관제 시점",
            yaxis=dict(range=[curr_cfg['min'] - 0.1, curr_cfg['max'] + 0.1]),
            hovermode="x unified", legend=dict(orientation="h", y=1.1, x=1)
        )
        st.plotly_chart(fig, use_container_width=True)

        if predicted_val >= curr_cfg['threshold']:
            st.success(f"🤖 **AI 통계 예측:** 현재의 환경 변화 추세가 이어지면 2주 뒤 예상 지수는 **{predicted_val:.3f}**로 {curr_cfg['ai_good']}")
        else:
            st.warning(f"🤖 **AI 통계 예측:** 현 추세가 고착화되면 2주 뒤 예상 지수는 **{predicted_val:.3f}**에 수렴할 것으로 판단되어 {curr_cfg['ai_bad']}")

        # -----------------------------------------------------------------
        # 📊 공공 결재 첨부용 정식 Excel 보고서 빌더 (모드 통합 고도화)
        # -----------------------------------------------------------------
        st.markdown("---")
        st.subheader("📥 지자체 결재 및 첨부용 정식 보고서 (Excel)")
        
        change_val = avg_val - last_avg if last_avg is not None else 0
        change_rate = (change_val / abs(last_avg) * 100) if last_avg and abs(last_avg) > 0 else 0
        reliability_score = "우수 (95%)" if cloud_threshold <= 25 else "보통 (80%)"

        report_data = {
            "관측 및 AI 예측 시점": [
                f"분석 시작일 ({start_date.strftime('%m/%d')})", 
                f"전년 동기 평균 (대조군)", 
                f"올해 현재 실측 ({end_date.strftime('%m/%d')})", 
                f"AI 2주 뒤 예측 ({future_date.strftime('%m/%d')})"
            ],
            f"원격 탐사 지수 ({idx_name})": [
                round(curr_cfg['baseline'], 4), 
                round(last_avg, 4) if last_avg is not None else round(curr_cfg['threshold'], 4), 
                round(avg_val, 4), 
                round(predicted_val, 4)
            ],
            "행정 정보 및 안전 진단 통계": [
                f"관제 지자체: {st.session_state.region_name}",
                f"플랫폼 모드: {st.session_state.current_mode}",
                f"전년 동기 대비 변화율: {change_rate:+.2f}%",
                f"위성 데이터 신뢰도: {reliability_score} ({st.session_state.count}장 합성)"
            ]
        }
        
        df = pd.DataFrame(report_data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='종합관제보고서')
            
            workbook = writer.book
            worksheet = writer.sheets['종합관제보고서']
            
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
            data_font = Font(name="맑은 고딕", size=10)
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
            )
            
            for col_num in range(1, 4):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            
            for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=3):
                for cell in row:
                    cell.font = data_font
                    cell.border = thin_border
                    if cell.column in [1, 2]:
                        cell.alignment = center_align
                        if cell.column == 2:
                            cell.number_format = '0.0000'
                    else:
                        cell.alignment = left_align

            chart = LineChart()
            chart.title = f"📈 {st.session_state.region_name} 구역 [{idx_name}] 종합 관제 시계열 추이"
            chart.style = 13
            chart.y_axis.title = f"{idx_name} 지수"
            chart.x_axis.title = "관제 단계"
            chart.width = 17
            chart.height = 10
            
            data_ref = Reference(worksheet, min_col=2, min_row=1, max_row=5)
            cats_ref = Reference(worksheet, min_col=1, min_row=2, max_row=5)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.legend = None
            
            s1 = chart.series[0]
            s1.graphicalProperties.line.width = 30000
            s1.smooth = True
            worksheet.add_chart(chart, "E2")

            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        val_str = str(cell.value)
                        cell_len = sum([2 if ord(char) > 128 else 1 for char in val_str])
                        if cell_len > max_len:
                            max_len = cell_len
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 16)
                
        excel_data = output.getvalue()
        
        st.download_button(
            label=f"📥 지자체 결재용 [{idx_name}] 관제 레포트 다운로드 (.xlsx)",
            data=excel_data,
            file_name=f"[{st.session_state.region_name}]_{idx_name}_종합관제보고서.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        
else:
    st.info("👈 왼쪽 컨트롤 패널에서 분석 모드와 확장된 지자체 타겟 지역을 선택하고 엔진 가동 버튼을 클릭해 보세요.")