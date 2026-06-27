# report_builder.py
"""
Excel 보고서 생성 모듈.

변경 사항:
- generate_excel_report() 가 이제 AnalysisResult를 직접 받는다.
  기존: app.py에서 DataFrame을 미리 조립해서 넘겼다.
        → 보고서 데이터 구성 책임이 UI 레이어에 있었다.
  개선: DataFrame 조립까지 이 모듈이 담당한다.
- 타입 힌트 추가.
- 로깅 추가.
"""
from __future__ import annotations

import io
import logging

import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from models import AnalysisResult

logger = logging.getLogger(__name__)


def _build_report_dataframe(result: AnalysisResult) -> pd.DataFrame:
    """
    AnalysisResult → 보고서용 DataFrame 변환.

    기존: app.py의 with st.expander(...) 블록 안에서 인라인으로 조립됐다.
    개선: 데이터 조립 책임을 report_builder로 이동.
    """
    req = result.request
    cur = result.current
    ly = result.last_year
    idx_name = req.mode_key  # 짧은 별칭 대신 mode_key를 그대로 사용하거나 index_name으로 대체 가능

    change_rate_display = (
        f"{result.change_rate:+.2f}%" if result.change_rate is not None
        else "비교 불가 (전년 데이터 없음)"
    )

    observation_labels = [
        "전년 동기 평균 (대조군)" if ly.has_data else "전년 동기 데이터 없음",
        f"올해 실측 평균 ({req.end_date.strftime('%m/%d')})",
        "구역 내 분포 통계 (표준편차)",
    ]

    from mode_config import mode_config
    cfg = mode_config[req.mode_key]
    index_name = cfg["index_name"]

    index_values = [
        round(ly.mean, 4) if ly.has_data and ly.mean is not None else None,
        round(cur.mean, 4) if cur.mean is not None else None,
        round(cur.std_dev, 4) if cur.std_dev is not None else None,
    ]

    region_name = req.region.name
    mode_label = req.mode_key

    if cur.min_val is not None and cur.max_val is not None:
        dist_info = f"구역 내 최솟값: {cur.min_val:.4f} / 최댓값: {cur.max_val:.4f}"
    else:
        dist_info = "구역 내 최소/최댓값 데이터 없음"

    admin_info = [
        f"관제 지자체: {region_name} / 플랫폼 모드: {mode_label}",
        f"전년 동기 대비 변화율: {change_rate_display} / 위성 데이터 신뢰도: {result.reliability_label}",
        dist_info,
    ]

    return pd.DataFrame({
        "관측 시점": observation_labels,
        f"원격 탐사 지수 ({index_name})": index_values,
        "행정 정보 및 안전 진단 통계": admin_info,
    })


def generate_excel_report(result: AnalysisResult) -> bytes:
    """
    AnalysisResult를 받아 Excel 바이너리를 반환한다.

    기존 시그니처:
        generate_excel_report(df, idx_name, region_name, current_mode,
                              change_rate, reliability_score, count)
    개선 시그니처:
        generate_excel_report(result: AnalysisResult) → bytes
    → 인자 7개 → 1개. 호출부가 데이터를 조립해 넘기는 부담 제거.
    """
    from mode_config import mode_config
    cfg = mode_config[result.request.mode_key]
    idx_name = cfg["index_name"]
    region_name = result.request.region.name

    df = _build_report_dataframe(result)

    output = io.BytesIO()
    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="종합관제보고서")
            workbook = writer.book
            worksheet = writer.sheets["종합관제보고서"]

            # ── 스타일 정의 ──────────────────────────────────────────
            header_fill = PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid"
            )
            header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
            data_font = Font(name="맑은 고딕", size=10)
            center_align = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9"),
            )

            # ── 헤더 스타일 ──────────────────────────────────────────
            for col_num in range(1, 4):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            # ── 데이터 행 스타일 ─────────────────────────────────────
            for row in worksheet.iter_rows(
                min_row=2, max_row=len(df) + 1, min_col=1, max_col=3
            ):
                for cell in row:
                    cell.font = data_font
                    cell.border = thin_border
                    if cell.column in (1, 2):
                        cell.alignment = center_align
                        if cell.column == 2:
                            cell.number_format = "0.0000"
                    else:
                        cell.alignment = left_align

            # ── 라인 차트 ────────────────────────────────────────────
            chart = LineChart()
            chart.title = f"📈 {region_name} 구역 [{idx_name}] 종합 관제 시계열 추이"
            chart.style = 13
            chart.y_axis.title = f"{idx_name} 지수"
            chart.x_axis.title = "관제 단계"
            chart.width = 17
            chart.height = 10

            data_ref = Reference(worksheet, min_col=2, min_row=1, max_row=len(df) + 1)
            cats_ref = Reference(worksheet, min_col=1, min_row=2, max_row=len(df) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.legend = None

            series = chart.series[0]
            series.graphicalProperties.line.width = 30000
            series.smooth = True
            worksheet.add_chart(chart, "E2")

            # ── 컬럼 너비 자동 조정 ──────────────────────────────────
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        cell_len = sum(
                            2 if ord(char) > 128 else 1
                            for char in str(cell.value)
                        )
                        max_len = max(max_len, cell_len)
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 16)

    except Exception:
        logger.exception("Excel 보고서 생성 실패 | region=%s mode=%s", region_name, idx_name)
        raise

    return output.getvalue()
